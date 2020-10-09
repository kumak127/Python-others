#! python3
# xlsx_text_xlsx.py - xlsxファイルを列ごとのテキストファイルにする

import openpyxl

file_path = input(r"テキストファイルにするxlsxファイルのパスを入力してください:")

wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

# シートの列数を調べる
max_column = sheet.max_column

# シートの列ごとにテキストファイルを作り、書き込んでいく
for i in range(1, max_column+1):
    column = openpyxl.utils.get_column_letter(i)
    cell_obj_of_column = sheet[column]
    with open(r"..\pydata\xlsx_text_{}.txt".format(column), "w", encoding="utf-8") as f:
        for cell_obj in cell_obj_of_column:
            f.write(str(cell_obj.value) + "\n")

