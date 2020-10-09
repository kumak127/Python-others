#! python 
# blankRowInserter.py - 2つの整数とファイル名を受け取り1つ目の整数行から2つ目の整数行分の空行を加える

import openpyxl, sys
from os.path import basename
from openpyxl.utils import get_column_letter

if len(sys.argv) < 4:
    insert_num = int(input("空行を入れる行の場所を整数で指定してください:"))
    blank_row_num = int(input("空行を入れる行数を整数で指定してください:"))
    file_path = input(r"処理を行うファイル名のパスを指定してください")

else:
    insert_num = int(sys.argv[1])
    blank_row_num = int(sys.argv[2])
    file_path = sys.argv[3]

# 変更前のシートを読み込む
old_wb = openpyxl.load_workbook(file_path)
old_sheet = old_wb.active

# 新しく作るファイルとシートを作る
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# 指定された行より上の内容を新しいファイルに書き込む

for row_of_cell_obj in old_sheet[1: insert_num-1]:
    for cell_obj in row_of_cell_obj:
        new_sheet[cell_obj.coordinate].value = cell_obj.value

# 指定された行よりも下の内容を指定された行分下げて書き込む
for row_of_cell_obj in old_sheet[insert_num: old_sheet.max_row]:
    for cell_obj in row_of_cell_obj:
        new_sheet["{}{}".format(get_column_letter(cell_obj.column), cell_obj.row + blank_row_num)].value = cell_obj.value


new_wb.save(r"C:\Users\new_{}".format(basename(file_path)))
