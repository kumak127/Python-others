#! pyhton3
# Excel_to_csv.py - フォルダ内のすべてのxlsxファイルをcsvファイルに変換する

import openpyxl, csv, os

os.chdir(input(r"Excelファイルの存在するフォルダのパスを入力してください:"))

# カレントディレクトリのすべてのファイルを検出
for excel_file in os.listdir("."):
    if not excel_file.endswith(".xlsx"):                                    # 拡張子が.xlsx以外ならコンティニュー
        continue
    wb = openpyxl.load_workbook(excel_file)                                 # Workbookオブジェクトを作成
    excel_file_name = excel_file.split(".")[0]                              # 拡張子よりも前の名前を入手
    for sheet_name in wb.get_sheet_names():                                 # シート名のリストを取得して一つずつループ
        sheet = wb.get_sheet_by_name(sheet_name)                            # Sheetオブジェクトを作る
        csv_file_name = excel_file_name + "_" + sheet_name + ".csv"         # csvファイルの名前を付ける
        with open(csv_file_name,"w",newline="") as csv_file:                # csvファイルの作成し、書き込みで開く
            csv_writer = csv.writer(csv_file)                               # csvのWriterオブジェクトを作る
            for row_num in range(1,sheet.max_row+1):                        # 1～最大行までの数字をループ
                row_data = []                                               # 行のすべての値を入れる空のリスト作る
                for column_num in range(1,sheet.max_column+1):              # 1～最大列までの数字をループ
                    row_data.append(sheet.cell(row_num,column_num).value)   # セルの一つずつの値を行の値を入れるリストに入れていく
                csv_writer.writerow(row_data)                               # 行の値がすべて入ったら、csvファイルに書き込む

