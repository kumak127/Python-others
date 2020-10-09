#! python3
# changeRowColumn.py - エクセルファイルの行と列を入れ変える

import openpyxl, os
from openpyxl.utils import get_column_letter, column_index_from_string

file_path = input(r"ファイルのパスを入力してください")

# 変更する前のファイルとシートを開く
old_wb = openpyxl.load_workbook(file_path)
old_sheet = old_wb.active

# 変更後のファイルとシートを作る
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# 変更前のシートの行数と列数を調べる
row_num = old_sheet.max_row
column_num = old_sheet.max_column

# 変更前のセルの値を変更後のセルに行と列を入れ替えて入れていく
for row_of_cell_obj in old_sheet["A1": f"{get_column_letter(column_num)}{row_num}"]:
    for cell_obj in row_of_cell_obj:
        new_sheet.cell(cell_obj.column, cell_obj.row, cell_obj.value)

dire = os.path.dirname(file_path)
file_name = "new_" + os.path.basename(file_path)

new_wb.save(os.path.join(dire, file_name))