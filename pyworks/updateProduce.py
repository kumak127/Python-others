#! python3
# updateProduce.py - 農作物スプレッドシートを価格を訂正する

import openpyxl

wb = openpyxl.load_workbook(r"..\pydata\produceSales.xlsx")
sheet = wb.get_sheet_by_name("Sheet")

# 農作物の種類と更新する価格

PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

# 行をループして価格を更新する
for row_num in range(2,sheet.max_row + 1):  # 先頭行はスキップ
    produce_name = sheet.cell(row_num, column=1).value   # A列の文字列を入手
    if produce_name in PRICE_UPDATES:   # 入手した文字列が変更したい農作物のキーにあるなら
        sheet.cell(row_num, column=2).value = PRICE_UPDATES[produce_name]    # その行のB列のセルを、更新する価格に変更

wb.save(r"..\pydata\updateProduceSales.xlsx")