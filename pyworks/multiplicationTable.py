#! python3
# multiplicationTable.py - N×Nの掛け算の表を作る

import openpyxl, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

if len(sys.argv) < 2:
    num = int(input("作る掛け算の表の数値を入力してください"))

else:
    num = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active

# 表の1行目とA列目に数値を順番に入れていく(太字にする)
bold_font = Font(bold=True)
for i in range(1, num+1):
    sheet.cell(i+1, 1, i).font = bold_font

for i in range(1,num+1):
    sheet.cell(1, i+1, i).font = bold_font

# 表の一行目と一列目は固定する
sheet.freeze_panes = "B2"

# 表の内側にそれぞれの掛け算の結果を入れていく
for row_of_cell_obj in sheet["B2":"{}{}".format(get_column_letter(num+1), num+1)]:
    for cell_obj in row_of_cell_obj:
        cell_obj.value = "=A{}*{}1".format(cell_obj.row, get_column_letter(cell_obj.column))

wb.save(r"..\pydata\multiplicationTable{}.xlsx".format(num))